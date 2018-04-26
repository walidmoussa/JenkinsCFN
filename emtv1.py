# ============================================================================= #
#    Automating MASS Test Injection / Export reports and Upload Tool v1.1.0     #
#         Created by Walid MOUSSA 2016-02 for ESSILOR international             #
#                      Last update : April-12 2016                              #
#                           Using Python 2.7                                    #
# ============================================================================= #

#!/usr/bin/env python 

# Drive API Imports
from __future__ import print_function
from googleapiclient.discovery import build
from httplib2 import Http
from oauth2client import file, client, tools
from oauth2client.client import flow_from_clientsecrets
from oauth2client.file import Storage
import requests
import certifi

# Test Injection & GUI Imports
import pymssql
from ConfigParser import RawConfigParser
import _mssql
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
import Tkinter
from Tkinter import *
import ttk
import tkMessageBox
from tkMessageBox import *
import tkFileDialog
import Pmw
from Pmw import *
import re
import xlsxwriter

# Copy/Past Imports 
import glob
import shutil
import os
from os import rename,listdir
from os.path import basename
from re import sub
import re
import warnings
import datetime
from time import sleep, strftime, strptime
warnings.simplefilter("ignore")

global myGUI
class CopyBat():
    
    # Initilalize the template
    def __init__(self, master):
        self.parent = master
        self.gui()
    # Create the interface 
    def gui(self):
        self.Source=StringVar()
        self.Destination=StringVar()
        self.cpNbr=IntVar()
        self.cpTime= IntVar()
        
        # Label, Button ...
        label = Label(myGUI, text='Source folder').grid(row=9,column=0)
        MySource = Entry(myGUI, textvariable=self.Source, width=33).grid(row=9, column=2)
        browse = Button(myGUI,text="Browse",bg="lightseagreen", command=lambda:self.Source.set(tkFileDialog.askdirectory(initialdir='//frer0891')), cursor="hand2").grid(row=9, column=3)
        label = Label(myGUI, text='Target folder').grid(row=10,column=0)
        MyDestination=Entry(myGUI, textvariable=self.Destination, width=33).grid(row=10, column=2)
        browse1=Button(myGUI,text="Browse",bg="lightseagreen",command=lambda:self.Destination.set(tkFileDialog.askdirectory(mustexist = False)), cursor="hand2").grid(row=10, column=3)
        
        button2=Button(myGUI, text="  Archive  ", command=self.archiv,bg="bisque", cursor="hand2").grid(row=11, column=2)
        
        label = Label(myGUI, text='Files number').grid(row=12,column=0)
        cpnbr = Entry(myGUI, textvariable = self.cpNbr, width=5).grid(row=12, column=1)
        #label.place(relx= 0.22, rely= 0.55, anchor= NE)
        #cpnbr.place(relx= 0.34, rely= 0.55, anchor= NE)
        label = Label(myGUI, text='Delay (min)').grid(row=13,column=0)
        #label2.place(relx= 0.22, rely= 0.9, anchor= NE)
        cprange =Entry(myGUI, textvariable=self.cpTime, width=5).grid(row=13,column=1)
        #cprange.place(relx= 0.34, rely= .9, anchor= NE)
        
        button1=Button(myGUI, text="Start Batch", command=self.copyy,bg="aquamarine", cursor="hand2").grid(row=70, column=1)
        
        #label3 = Label(myGUI, text='Copy Status').grid(row=70,column=0)
        #cpstatus =Entry(myGUI, textvariable=self.cpStatus, width=35, bg ='whitesmoke', state = DISABLED ).grid(row=70, column=2)
        
    def archiv(self):
        
            self.parent.withdraw()
            infoMessage = "Archive is running..." 
            showinfo("Running...", infoMessage)
            global archv_dir
            archv_dir = (self.Destination.get())
            if not os.path.exists(archv_dir):
                os.makedirs(archv_dir)
            src_dir = (self.Source.get())
            nfl = len(os.listdir(src_dir))
            status = "--- %d Files successfully copied ---" % nfl
            for myfiles in glob.iglob(os.path.join(src_dir, "*.*")):
                shutil.copy(myfiles, archv_dir)
            self.parent.deiconify()
            showinfo("Finish", status)
        
    def copyy(self):
            self.parent.withdraw()
            #self.cpStatus.set('')
            infoMessage = "Scheduled batch start" 
            showinfo("Running...", infoMessage)
            dest_dir = (self.Destination.get())
            str_the_path = (self.Source.get())
            global nfl
            nfl = len(os.listdir(str_the_path)) 
            status = "--- %d Batchs successfully launched ---" % nfl
            #self.cpStatus.set(status)
            
            y = (self.cpNbr.get())
            tts = (self.cpTime.get() * 60)
            z = 3
            while True:
                x = 1
                for str_each_file in listdir(str_the_path):
                    # look for the files we want
                    if str_each_file.endswith('.xml'):
                            # get the basename (remove extension)
                            str_xml_name = sub('.xml','',str_each_file)
                            # now add the new one
                            str_new_name = str(x) + '-' + 'new' + str (z) + '.xml'
                            count = int(x)
                            x += 1
                            # full path for both files
                            str_old_name = '/'.join([str_the_path,str_each_file])
                            str_new_name = '/'.join([str_the_path,str_new_name])
                            # and the full path to the files
                            rename(str_old_name,str_new_name)
                            print (z)
                            
                for str_each_file in glob.iglob(os.path.join(str_the_path, "*.*")):
                        names = os.path.splitext(os.path.basename(str_each_file))[0]
                        b_min = 1
                        b_max = y + 1
                        for i in range(b_min, b_max):
                                j = str(i) + "-" 
                                if names.startswith(j):
                                        shutil.move(str_each_file, dest_dir)
                                        
                sleep(tts)
                z += 150
                if len(os.listdir(str_the_path)) == 0:
                    break
            
                
            self.parent.deiconify()
            infoMessage = "Scheduled batch finish \n%s" % status 
            showinfo("Finish", infoMessage)
            
class QueryWindow(Frame):
    """GUI Database Query Frame"""
    def __init__(self):
        """QueryWindow Constructor"""
        Frame.__init__(self)
        Pmw.initialise()
        self.pack(expand = YES, fill = BOTH)
        self.master.title("Essilor MASS Test v1.1.5")
        self.master.geometry("600x320")
        self.master.resizable(width = FALSE, height = FALSE)
        
        # Create and pack OptionMenu for Instance (server)
        self.var = Tkinter.StringVar()
        self.var.set('IE301')
        self.method_menu = Pmw.OptionMenu(self,
                labelpos = 'w',
                label_text = 'Choose instance :',
                menubutton_textvariable = self.var,
                items = ['IE301', 'IE401', 'IE402', 'IE501', 'IFF02', 'New server'],
                menubutton_width = 15,
        )
        self.method_menu.pack(anchor = 'n')
        
        # Buttons to browse and add new instance file 
        self.browse1 = Button(self, text ="Add (*.ini)", height = 1, width = 9, command = self.browseInstance, bg ="lightblue", cursor="hand2")
        #self.browse1.pack(pady = 7)
        self.browse1.place(relx= .83, rely= .006, anchor= NE)
           
        # Buttons to browse and update original SQL file 
        self.browse = Button(self, text ="Step 1: Click to browse SQL file", command = self.browseQuery, bg ="lightblue", cursor="hand2")
        self.browse.pack(fill = X, padx = 17, pady = 7)
        
        self.queryUpdate1 = Pmw.EntryField (self, labelpos = 'w', label_text = '1st DateTime :       ', value = 'YYYY-MM-DD hh:mm', validate = None)
        self.queryUpdate1.pack(fill = X, padx = 17, pady = 4)
        
        self.queryUpdate2 = Pmw.EntryField (self, labelpos = 'w', label_text = '2nd DateTime :     ', value = 'YYYY-MM-DD hh:mm', validate = None)
        self.queryUpdate2.pack(fill = X, padx = 17, pady = 4)
        
        self.queryUpdate3 = Pmw.EntryField (self, labelpos = 'w', label_text = 'Entity ID (Order DB Test) :                   ', value = 'XXX_XX' , validate = None)
        self.queryUpdate3.pack(fill = X, padx = 17, pady = 4)
        
        #self.queryUpdate4 = Pmw.EntryField (self, labelpos = 'w', label_text = 'Delivery Unit ID (On Route DB Test) :', value = 'YYYYMMDD', validate = None, entry_background = 'whitesmoke', entry_state = DISABLED)
        #self.queryUpdate4.pack(fill = X, padx = 17, pady = 4)
        
        # Submit Query button and status label
        self.submit = Button(self, text = "Step 2: Click to submit test and export results", command = self.submitQuery, bg = "mediumseagreen", cursor = "hand2")
        self.submit.pack(fill = X, padx = 17, pady = 7)
        self.submitStatus = Pmw.EntryField (self, labelpos = 'w', label_text = 'Query Status :   ', entry_background = 'whitesmoke', validate = None, entry_state = DISABLED)
        self.submitStatus.pack(fill = X, padx = 17, pady = 7)
        
        # Button to upload to Drive and status label 
        self.upload = Button(self, text ="Step 3: Click to upload to your Drive", command = self.uploadDrive, bg = "salmon", cursor = "hand2")
        self.upload.pack(fill = X, padx = 17, pady = 7)
        self.uploadStatus = Pmw.EntryField (self, labelpos = 'w', label_text = 'Upload Status : ', entry_background = 'whitesmoke', validate = None, entry_state = DISABLED)
        self.uploadStatus.pack(fill = X, padx = 17, pady = 7)
         
    """Database Query Submit & Export results to Excel file"""
    
    def browseInstance(self):
        if self.var.get() == 'New server':
            global file_path2
            file_path2 = tkFileDialog.askopenfilename()
            global fp2
            global fpO1
            try:
                global fins
                fins = open(file_path2, 'r')
            except:
                errorMessage = "Please select a file\nOnly \"ini\" format accepted"
                showerror("Error", errorMessage)
            
    def browseQuery(self):
        
        # Refresh window for new input-file
        self.submitStatus.setentry('')
        self.uploadStatus.setentry('')
        
        # Auto-complete values
        crtm = time.strftime("%Y-%m-%d %H:%M")
        lastd = str(datetime.date.today()-datetime.timedelta(1))
        lastdtm = lastd + " 00:00"
        self.queryUpdate1.setentry(lastdtm)
        self.queryUpdate2.setentry(crtm)
        
        if self.var.get() == 'IE301':
            self.queryUpdate3.setentry('')
            self.queryUpdate3.setentry('\'ESS_DE\',\'ESS_AT\'')
            
        elif self.var.get() == 'IE401':
            self.queryUpdate3.setentry('')
            self.queryUpdate3.setentry('\'ESS_NL\',\'ESS_BE\',\'ESS_DK\',\'ESS_FI\',\'ESS_SW\',\'ESS_NO\'')
            
        elif self.var.get() == 'IE402':
            self.queryUpdate3.setentry('')
            self.queryUpdate3.setentry('\'ESS_FI\',\'ESS_SW\',\'ESS_NO\'')
            
        elif self.var.get() == 'IE501':
            self.queryUpdate3.setentry('')
            self.queryUpdate3.setentry('\'ESS_DE\',\'ESS_AT\'')
            
        elif self.var.get() == 'IFF02':
            self.queryUpdate3.setentry('')
            self.queryUpdate3.setentry('\'ESS_FF\'')
                
        file_path1 = tkFileDialog.askopenfilename()
        global fp2
        global fpO1
        try:
            fp2 = open(file_path1, 'r')
        except:
            errorMessage = "Please select a file\nOnly SQL format accepted"
            showerror("Error", errorMessage)    
    
    def submitQuery(self):
        """Update & Execute user-entered query file again database"""
        
        # Update query entries
        fpO1 = open('C:/Users/toExecute.sql',"w")
        data = fp2.read()
        fp2.close()
        nvlVal1 = self.queryUpdate1.get()
        nvlVal2 = self.queryUpdate2.get()
        nvlVal3 = self.queryUpdate3.get()
        nvlVal4 = nvlVal2[:4] + nvlVal2[5:7] + nvlVal2[8:10]
        print (nvlVal4)
        mydict = {'myNV1': nvlVal1, 'myNV2': nvlVal2, 'myNV3' : nvlVal3, 'myNV4' : nvlVal4}
            
        for key, value in mydict.items():
            data = data.replace(key, value)  
        fpO1.write(data)
        fpO1.close()
                      
        # Prepare the file and the query
        fpO1 = open("C:/Users/toExecute.sql")
        outFile1 = open("C:/Users/query1.sql", "w")
        outFile2 = open("C:/Users/query2.sql", "w")
        outFile3 = open("C:/Users/query3.sql", "w")
        outFile4 = open("C:/Users/query4.sql", "w")
        buffer = []
        keepCurrentSet = True
        for line in fpO1:
            buffer.append(line)
            if line.startswith("/*A"):
                #---- starts a new data set
                if keepCurrentSet:
                    outFile1.write("".join(buffer))
                #now reset our state
                keepCurrentSet = False
                buffer = []
        
            elif line.startswith("/*B"):
                if keepCurrentSet:
                    outFile2.write("".join(buffer))
                keepCurrentSet = False
                buffer = []
        
            elif line.startswith("/*C"):
                if keepCurrentSet:
                    outFile3.write("".join(buffer))
                keepCurrentSet = False
                buffer = []
                
            elif line.startswith("/*D"):
                if keepCurrentSet:
                    outFile4.write("".join(buffer))
                keepCurrentSet = False
                buffer = []
        
            elif line.startswith(";"):
                keepCurrentSet = True
        
        fpO1.close()
        outFile1.close()
        outFile2.close()
        outFile3.close()
        outFile4.close()
        
        outFile1 = open("C:/Users/query1.sql", "r")
        qury1 = outFile1.read()
        
        outFile2 = open("C:/Users/query2.sql", "r")
        qury2 = outFile2.read()
        
        outFile3 = open("C:/Users/query3.sql", "r")
        qury3 = outFile3.read()
        
        outFile4 = open('C:/Users/query4.sql', 'r')
        qury4 = outFile4.read()
        
        status = ' --- Connection successful --- Query successfully executed ---'
        global cursor1
        global cursor2
        global cursor3
        global cursor4
        global data1
        global data2
        global data3
        global data4
        global fields1, fields2, fields3, fields4
        global names1, names2, names3, traiames4
            
        # Select instance, Open connection, retrieve cursor and execute query
        if self.var.get() == 'IE301':
            connection = pymssql.connect(server ='FRER0973\IO13',user = 'OMUser', password = '%Essilor1%', database = 'OrderDB')
            connection2 = pymssql.connect(server ='FRER0973\ION3',user = 'OMUser', password = '%Essilor1%', database = 'OnRouteDB')
            
            cursor1 = connection.cursor()
            cursor1.execute(qury1)
            data1 = cursor1.fetchall()
            fields1 = cursor1.description # metadata from query
            names1 = [x[0] for x in cursor1.description]
            cursor1.close()
            
            cursor2 = connection.cursor()
            cursor2.execute(qury2)
            data2 = cursor2.fetchall()
            fields2 = cursor2.description # metadata from query
            names2 = [x[0] for x in cursor2.description]
            cursor2.close()
            
            cursor3 = connection.cursor()
            cursor3.execute(qury3)
            data3 = cursor3.fetchall()
            fields3 = cursor3.description # metadata from query
            names3 = [x[0] for x in cursor3.description]
            cursor3.close()
            
            cursor4 = connection2.cursor()
            cursor4.execute(qury4)
            data4 = cursor4.fetchall()
            fields4 = cursor4.description # metadata from query
            names4 = [x[0] for x in cursor4.description]
            cursor4.close()
            
            self.submitStatus.setentry(status)
                
        elif self.var.get() == 'IE401':
            connection = pymssql.connect(server='FRER0972\IO14', user='OMUser', password='%Essilor1%', database='OrderDB')
            connection2 = pymssql.connect(server='FRER0972\ION4', user='OMUser', password='%Essilor1%', database='OnRouteDB')
            
            cursor1 = connection.cursor()
            cursor1.execute(qury1)
            data1 = cursor1.fetchall()
            fields1 = cursor1.description # metadata from query
            names1 = [x[0] for x in cursor1.description]
            cursor1.close()
            
            cursor2 = connection.cursor()
            cursor2.execute(qury2)
            data2 = cursor2.fetchall()
            fields2 = cursor2.description # metadata from query
            names2 = [x[0] for x in cursor2.description]
            cursor2.close()
            
            cursor3 = connection.cursor()
            cursor3.execute(qury3)
            data3 = cursor3.fetchall()
            fields3 = cursor3.description # metadata from query
            names3 = [x[0] for x in cursor3.description]
            cursor3.close()
            
            cursor4 = connection2.cursor()
            cursor4.execute(qury4)
            data4 = cursor4.fetchall()
            fields4 = cursor4.description # metadata from query
            names4 = [x[0] for x in cursor4.description]
            cursor4.close()
            
            self.submitStatus.setentry(status)
                
        elif self.var.get() == 'IE402':
            connection = pymssql.connect(server='FRER1420\IO14', user='OMUser', password='%Essilor1%', database='OrderDB')
            connection2 = pymssql.connect(server='FRER1420\ION4', user='OMUser', password='%Essilor1%', database='OnRouteDB')
            
            cursor1 = connection.cursor()
            cursor1.execute(qury1)
            data1 = cursor1.fetchall()
            fields1 = cursor1.description # metadata from query
            names1 = [x[0] for x in cursor1.description]
            cursor1.close()
            
            cursor2 = connection.cursor()
            cursor2.execute(qury2)
            data2 = cursor2.fetchall()
            fields2 = cursor2.description # metadata from query
            names2 = [x[0] for x in cursor2.description]
            cursor2.close()
            
            cursor3 = connection.cursor()
            cursor3.execute(qury3)
            data3 = cursor3.fetchall()
            fields3 = cursor3.description # metadata from query
            names3 = [x[0] for x in cursor3.description]
            cursor3.close()
            
            cursor4 = connection2.cursor()
            cursor4.execute(qury4)
            data4 = cursor4.fetchall()
            fields4 = cursor4.description # metadata from query
            names4 = [x[0] for x in cursor4.description]
            cursor4.close()
            
            self.submitStatus.setentry(status)
                
        elif self.var.get() == 'IE501':
            connection = pymssql.connect(server='FRER1308\IO14', user='OMUser', password='%Essilor1%', database='OrderDB')
            connection2 = pymssql.connect(server='FRER1308\ION4', user='OMUser', password='%Essilor1%', database='OnRouteDB')
            
            cursor1 = connection.cursor()
            cursor1.execute(qury1)
            data1 = cursor1.fetchall()
            fields1 = cursor1.description # metadata from query
            names1 = [x[0] for x in cursor1.description]
            cursor1.close()
            
            cursor2 = connection.cursor()
            cursor2.execute(qury2)
            data2 = cursor2.fetchall()
            fields2 = cursor2.description # metadata from query
            names2 = [x[0] for x in cursor2.description]
            cursor2.close()
            
            cursor3 = connection.cursor()
            cursor3.execute(qury3)
            data3 = cursor3.fetchall()
            fields3 = cursor3.description # metadata from query
            names3 = [x[0] for x in cursor3.description]
            cursor3.close()
            
            cursor4 = connection2.cursor()
            cursor4.execute(qury4)
            data4 = cursor4.fetchall()
            fields4 = cursor4.description # metadata from query
            names4 = [x[0] for x in cursor4.description]
            cursor4.close()
            
            self.submitStatus.setentry(status)
                
        elif self.var.get() == 'IFF02':
            connection = pymssql.connect(server='FRER1344\IO1F', user='OMUser', password='%Essilor1%', database='OrderDB')
            connection2 = pymssql.connect(server='FRER1344\IONF', user='OMUser', password='%Essilor1%', database='OnRouteDB')
            
            cursor1 = connection.cursor()
            cursor1.execute(qury1)
            data1 = cursor1.fetchall()
            fields1 = cursor1.description # metadata from query
            names1 = [x[0] for x in cursor1.description]
            cursor1.close()
            
            cursor2 = connection.cursor()
            cursor2.execute(qury2)
            data2 = cursor2.fetchall()
            fields2 = cursor2.description # metadata from query
            names2 = [x[0] for x in cursor2.description]
            cursor2.close()
            
            cursor3 = connection.cursor()
            cursor3.execute(qury3)
            data3 = cursor3.fetchall()
            fields3 = cursor3.description # metadata from query
            names3 = [x[0] for x in cursor3.description]
            cursor3.close()
            
            cursor4 = connection2.cursor()
            cursor4.execute(qury4)
            data4 = cursor4.fetchall()
            fields4 = cursor4.description # metadata from query
            names4 = [x[0] for x in cursor4.description]
            cursor4.close()
            
            self.submitStatus.setentry(status)
            
        elif self.var.get() == 'New server':
            
            parser = RawConfigParser()
            parser.read(file_path2)
            myserver = parser.get('add_instance', 'server')
            myserver2 = parser.get('add_instance', 'server2')
            myuser = parser.get('add_instance', 'user')
            mypassword = parser.get('add_instance', 'password')
            mydatabase = parser.get('add_instance', 'dbname')
            
            connection = pymssql.connect(server = myserver, user = myuser, password = mypassword, database = mydatabase)
            connection2 = pymssql.connect(server = myserver2, user = myuser, password = mypassword, database = mydatabase2)
            
            cursor1 = connection.cursor()
            cursor1.execute(qury1)
            data1 = cursor1.fetchall()
            fields1 = cursor1.description # metadata from query
            names1 = [x[0] for x in cursor1.description]
            cursor1.close()
            
            cursor2 = connection.cursor()
            cursor2.execute(qury2)
            data2 = cursor2.fetchall()
            fields2 = cursor2.description # metadata from query
            names2 = [x[0] for x in cursor2.description]
            cursor2.close()
            
            cursor3 = connection.cursor()
            cursor3.execute(qury3)
            data3 = cursor3.fetchall()
            fields3 = cursor3.description # metadata from query
            names3 = [x[0] for x in cursor3.description]
            cursor3.close()
            
            cursor4 = connection2.cursor()
            cursor4.execute(qury4)
            data4 = cursor4.fetchall()
            fields4 = cursor4.description # metadata from query
            names4 = [x[0] for x in cursor4.description]
            cursor4.close()
            
            self.submitStatus.setentry(status)
                
        else:
            status = '--- Connection failed --- No query executed ---'
            self.submitStatus.setentry(status)
          
        self.master.withdraw()
        infoMessage = "Please wait ! The export is running...\nThe window will appear once exporting finish.\nYou can continue to use your PC."
        showinfo("Running...", infoMessage)
            
        """Export query results to Excel file"""    
        df = pd.DataFrame(list(data1), columns = names1)
        df2 = pd.DataFrame(list(data2), columns = names2)
        df3 = pd.DataFrame(list(data3), columns = names3)
        df4 = pd.DataFrame(list(data4), columns = names4)
        
        # Save a copy of the xls file for upload
        save_path = 'C:/temp/Results.xlsx' 
        writer = pd.ExcelWriter(save_path, engine='xlsxwriter')
        list_dfs = [df, df2, df3, df4]
        for n, df in enumerate(list_dfs):
            df.to_excel(writer,'sheet%s' % n)
        writer.save()
        
        src_filename = 'C:/temp/Results.xlsx'
        dest_filename = 'C:/temp/Template_Report.xlsx'
        
        wb1 = openpyxl.load_workbook(dest_filename)
        wb2 = openpyxl.load_workbook(src_filename)
        
        ws2 = wb2.get_sheet_by_name('sheet0')
        ws22 = wb2.get_sheet_by_name('sheet1')
        ws222 = wb2.get_sheet_by_name('sheet2')
        ws2222 = wb2.get_sheet_by_name('sheet3')
        
        sh11_count = (ws222.max_row - 1)
        
        #ws0 = wb1.worksheets[0]
        ws1 = wb1.worksheets[1]
        ws11 = wb1.worksheets[2]
        ws111 = wb1.worksheets[3]
        ws1111 = wb1.worksheets[4]
        
        sh0_count = ws2.max_column
        sh1_count = ws22.max_column
        sh2_count = ws222.max_column
        sh3_count = ws2222.max_column
            
        for i in range(1, sh0_count):
            col_d = ws2.columns[i] # 0-indexing
            for idx, cell in enumerate(col_d, 1):
                ws1.cell(row = idx, column = i).value = cell.value #1-indexing
        ws1['A1'] = 'Request ID'
        ws1['B1'] = 'Order Entry'
        ws1['C1'] = 'Customer Number'
        ws1['D1'] = 'DN Number'
        
        ws1['P1'] = 'Creation Hour'
        ws1['Q1'] = 'Allocation Hour'
        ws1['R1'] = 'Diff Crea/Allo (in Second)'
        ws1['S1'] = 'Round Allocation Time (in Minute)'
            
        for j in range(1, sh1_count):
            col_d = ws22.columns[j] # 0-indexing
            for idx, cell in enumerate(col_d, 1):
                ws11.cell(row = idx, column = j).value = cell.value #1-indexing
        ws11['A1'] = 'Order Entry'
        ws11['B1'] = 'Customer Number'
        ws11['C1'] = 'Request ID'
        ws11['D1'] = 'DN Number'
        ws11['E1'] = 'Product Code'
        ws11['F1'] = 'Product Label'
        ws11['G1'] = 'Status'
        ws11['H1'] = 'Reason'
        ws11['I1'] = 'Lab'
        ws11['J1'] = 'Creation Date'
        ws11['K1'] = 'Severity'
        ws11['L1'] = 'Error Message'
        ws11['M1'] = 'Detailed Error Message'
            
        for k in range(1, sh2_count):
            col_d = ws222.columns[k] # 0-indexing
            for idx, cell in enumerate(col_d, 1):
                ws111.cell(row = idx, column = k).value = cell.value #1-indexing
        ws111['A1'] = 'Order Entry'
        ws111['B1'] = 'Customer Number'
        ws111['C1'] = 'Request ID'
        ws111['D1'] = 'DN Number'
        ws111['G1'] = 'Customer Reference'
        ws111['H1'] = 'Status'
        ws111['I1'] = 'Reason'
        ws111['J1'] = 'Lab'
        ws111['K1'] = 'Creation Date'
                    
        for l in range(1, sh3_count):
            col_d = ws2222.columns[l] # 0-indexing
            for idx, cell in enumerate(col_d, 1):
                ws1111.cell(row = idx, column = l).value = cell.value #1-indexing
        ws1111['B1'] = 'Request ID'
        ws1111['C1'] = 'Error Message'
        ws1111['D1'] = 'Error Value'
        
        
        '''for rowNumP in range(2, (ws1.get_highest_row()) + 1):
            if ws1.cell(row = rowNumP, column = 10).value is not None:
                CreationDT = ws1.cell(row = rowNumP, column = 10).value
                CreationT = str(CreationDT)
                Cre = re.search(' (.+?)\.', CreationT).group(1)
                CreT = datetime.datetime.strptime(Cre, "%H:%M:%S")
                CreationTime = CreT.strftime("%I:%M:%S %p")
                ws1.cell(row = rowNumP, column = 16).value = CreationTime
                creTime = time.strptime(CreationTime.split(' ')[0],'%H:%M:%S')
                CrSec = datetime.timedelta(hours=creTime.tm_hour,minutes=creTime.tm_min,seconds=creTime.tm_sec).total_seconds()
                
                if ws1.cell(row = rowNumP, column = 11).value is not None:
                    AllocationDT = ws1.cell(row = rowNumP, column = 11).value
                    AllocationT = str(AllocationDT)
                    Alo = re.search('(?<= )\S+', AllocationT).group(0)
                    OK = re.search(r'\.', Alo)
                    if OK:
                        AloT = datetime.datetime.strptime(Alo, "%H:%M:%S.%f")
                        AllocationTime = AloT.strftime("%I:%M:%S %p")
                        ws1.cell(row = rowNumP, column = 17).value = AllocationTime
                        aloTime = time.strptime(AllocationTime.split(' ')[0],'%H:%M:%S')
                        alSec = datetime.timedelta(hours=aloTime.tm_hour,minutes=aloTime.tm_min,seconds=aloTime.tm_sec).total_seconds()
                        ws1.cell(row = rowNumP, column = 18).value = alSec - CrSec
                        ws1.cell(row = rowNumP, column = 19).value = round(((ws1.cell(row = rowNumP, column = 18).value)/60), 1)
                        
                    else:
                        AloT = datetime.datetime.strptime(Alo, "%H:%M:%S")
                        AllocationTime = AloT.strftime("%I:%M:%S %p")
                        ws1.cell(row = rowNumP, column = 17).value = AllocationTime
                        aloTime = time.strptime(AllocationTime.split(' ')[0],'%H:%M:%S')
                        alSec = datetime.timedelta(hours=aloTime.tm_hour,minutes=aloTime.tm_min,seconds=aloTime.tm_sec).total_seconds()
                        ws1.cell(row = rowNumP, column = 18).value = alSec - CrSec
                        ws1.cell(row = rowNumP, column = 19).value = round(((ws1.cell(row = rowNumP, column = 18).value)/60), 1)
                        
                                   
            if ws1.cell(row = rowNumP, column = 1).value is not None and ws1.cell(row = rowNumP, column = 19).value is None:
                ws1.cell(row = rowNumP, column = 19).value = 0.00'''
                 
        wb1.save('C:/temp/ReportUP.xlsx')
    
        # Show the tool window again 
        status = "--- Successfully exported to the final reports file --- Look at C:\Users ---"
        self.submitStatus.setentry('')
        self.submitStatus.setentry(status)
        self.master.deiconify()
        infoMessage = "Allocation rate = %d commands" % sh11_count
        showinfo("Finish", infoMessage)        
        
    """Upload To Drive"""
    def uploadDrive(self):
        
        """Drive Connection"""
        try:
            import argparse
            flags = argparse.ArgumentParser(parents = [ tools.argparser ]).parse_args()
        except ImportError:
            flags = None

        SCOPES = 'https://www.googleapis.com/auth/drive.file'
        store = file.Storage('C:/tmp/storage.json')
        creds = store.get()
        if not creds or creds.invalid:
            flow = client.flow_from_clientsecrets('C:/tmp/client_secret.json', scope = SCOPES)
            creds = tools.run_flow(flow, store, flags) if flags else tools.run(flow, store)
        DRIVE = build('drive', 'v2', http = creds.authorize(Http(ca_certs = 'C:/tmp/cacert.pem')))
        
        xxx = 'C:/temp/ReportUP.xlsx'

        """Require The File & Upload"""
        FILES = (
        (xxx , True),   
        )
        try:
            
            for filename, convert in FILES:
                metadata = {'title': filename}
                res = DRIVE.files().insert(convert = convert, body = metadata, media_body = filename, fields = 'mimeType, exportLinks').execute()
        except:
            
                if  os.path.exists("C:/Users/query1.sql"):
                    os.remove("C:/Users/query1.sql")
                if  os.path.exists("C:/Users/query2.sql"):
                    os.remove("C:/Users/query2.sql")
                if  os.path.exists("C:/Users/query3.sql"):
                    os.remove("C:/Users/query3.sql")
                if  os.path.exists("C:/Users/query4.sql"):
                    os.remove("C:/Users/query4.sql")
                if  os.path.exists("C:/Users/toExecute.sql"):
                    os.remove("C:/Users/toExecute.sql")
                if  os.path.exists("C:/temp/Results.xlsx"):
                    os.remove("C:/temp/Results.xlsx")
                status = "--- Successfully uploaded to Drive ---"
                self.uploadStatus.setentry(status)

def main():
    QueryWindow().mainloop()
        
if __name__ == "__main__":
    myGUI = Tk()
    app = CopyBat(myGUI)
    myGUI.resizable(width = FALSE, height = FALSE)
    myGUI.title('Essilor CopyBat v1.1.5')
    myGUI.mainloop()
    main()